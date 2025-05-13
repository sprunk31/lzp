import base64
import time
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font

# Pagina instellingen
st.set_page_config(page_title="LZP Vergelijktool", page_icon="📊", layout="centered")

# ✅ Logo laden vanuit assets/logo.png
with open("assets/logo.png", "rb") as image_file:
    encoded = base64.b64encode(image_file.read()).decode()
    logo_html = f'''
        <div style="text-align: center; margin-bottom: 1rem;">
            <img src="data:image/png;base64,{encoded}" width="300">
        </div>
    '''

# ✅ Styling en logo tonen
st.markdown("""
    <style>
        .section-header {
            font-size: 1.3em;
            margin-top: 2rem;
            color: #34495e;
        }
        .stButton>button {
            background-color: #2c3e50;
            color: white;
            border-radius: 8px;
            padding: 0.5em 1em;
        }
        .stDownloadButton>button {
            background-color: #27ae60;
            color: white;
            border-radius: 8px;
            padding: 0.5em 1em;
        }
    </style>
""", unsafe_allow_html=True)

st.markdown(logo_html, unsafe_allow_html=True)

# 🔐 Login met gebruikers uit Streamlit secrets
gebruikers = st.secrets.get("auth", {})

if "ingelogd" not in st.session_state:
    st.session_state["ingelogd"] = False

def login():
    username = st.text_input("Gebruikersnaam")
    password = st.text_input("Wachtwoord", type="password")
    if st.button("Inloggen"):
        if gebruikers.get(username) == password:
            st.session_state["ingelogd"] = True
            st.success(f"✅ Ingelogd als {username}")
            st.experimental_rerun()
        else:
            st.error("❌ Ongeldige inloggegevens")

if not st.session_state["ingelogd"]:
    login()
    st.stop()

# 📁 Upload twee bestanden
st.markdown("<div class='section-header'>📂 Upload je Excelbestanden</div>", unsafe_allow_html=True)
prezero_file = st.file_uploader("Upload PreZero Excelbestand (.xlsm, .xlsx, .xls)", type=["xlsm", "xlsx", "xls"], key="prezero")
avalex_file = st.file_uploader("Upload Avalex Excelbestand (.xlsm, .xlsx, .xls)", type=["xlsm", "xlsx", "xls"], key="avalex")

if prezero_file and avalex_file:
    # Start timer
    start_time = time.perf_counter()
    try:
        prezero_sheets = pd.read_excel(prezero_file, sheet_name=None)
        avalex_sheets = pd.read_excel(avalex_file, sheet_name=None)
    except Exception as e:
        st.error(f"❌ Fout bij het lezen van de Excelbestanden: {e}")
        st.stop()

    # Automatisch kiezen of dropdown tonen
    if len(prezero_sheets) == 1:
        sheet_prezero = list(prezero_sheets.keys())[0]
    else:
        sheet_prezero = st.selectbox("Kies PreZero-sheet", options=list(prezero_sheets.keys()))

    if len(avalex_sheets) == 1:
        sheet_avalex = list(avalex_sheets.keys())[0]
    else:
        sheet_avalex = st.selectbox("Kies Avalex-sheet", options=list(avalex_sheets.keys()))

    df_prezero = prezero_sheets[sheet_prezero]
    df_avalex = avalex_sheets[sheet_avalex]

    waarde = "Suez Recycling Services Berkel"

    df_avalex_filtered = df_avalex[df_avalex['Bestemming'] == waarde].copy()
    df_avalex_rest = df_avalex[df_avalex['Bestemming'] != waarde].copy()

    required_prezero = ['weegbonnr', 'gewicht']
    required_avalex = ['Weegbonnummer', 'Gewicht(kg)']
    if all(col in df_prezero.columns for col in required_prezero) and all(col in df_avalex_filtered.columns for col in required_avalex):

        def normalize_bon(val):
            try:
                if pd.isna(val) or str(val).strip() == "":
                    return ""
                return str(int(float(val)))
            except:
                return str(val).strip()

        df_avalex_filtered['Weegbonnummer_genorm'] = df_avalex_filtered['Weegbonnummer'].apply(normalize_bon)
        df_prezero['weegbonnr_genorm'] = df_prezero['weegbonnr'].apply(normalize_bon)

        bon_dict = df_prezero.set_index('weegbonnr_genorm')['gewicht'].to_dict()

        resultaten = []
        redenen = []
        for _, row in df_avalex_filtered.iterrows():
            bon = row['Weegbonnummer_genorm']
            gewicht = row['Gewicht(kg)']
            if bon == "":
                resultaten.append("Geen bon aanwezig")
                redenen.append("")
            elif bon in bon_dict:
                gewicht_ref = bon_dict[bon]
                if pd.isna(gewicht) or round(gewicht, 1) != round(gewicht_ref, 1):
                    resultaten.append("Bon aanwezig")
                    redenen.append(gewicht_ref)
                else:
                    resultaten.append("Bon aanwezig")
                    redenen.append("")
            else:
                resultaten.append("Geen bon aanwezig")
                redenen.append("")

        df_avalex_filtered['komt voor in PreZero'] = resultaten
        df_avalex_filtered['Reden'] = redenen
        df_avalex_rest['komt voor in PreZero'] = ""
        df_avalex_rest['Reden'] = ""
        df_avalex_combined = pd.concat([df_avalex_filtered, df_avalex_rest], ignore_index=True)

        avalex_bonnen = df_avalex_combined['Weegbonnummer'].dropna().apply(normalize_bon).tolist()
        ontbrekende_mask = ~df_prezero['weegbonnr_genorm'].isin(avalex_bonnen)

        # Schrijven naar Excel met kleuring
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_prezero.drop(columns=['weegbonnr_genorm'], errors='ignore').to_excel(writer, sheet_name='PreZero', index=False)
            df_avalex_combined.drop(columns=['Weegbonnummer_genorm'], errors='ignore').to_excel(writer, sheet_name='Avalex', index=False)

            wb = writer.book
            ws_avalex = wb['Avalex']
            fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            font_white = Font(color="FFFFFF")

            for idx, cell in enumerate(ws_avalex[1], start=1):
                if cell.value == 'komt voor in PreZero': col_komt = idx
                if cell.value == 'Reden': col_red = idx

            for row in ws_avalex.iter_rows(min_row=2):
                if col_red and row[col_red-1].value not in (None, "", " "):
                    if col_komt:
                        cell = row[col_komt-1]
                        cell.fill = fill_red
                        cell.font = font_white

            ws_prezero = wb['PreZero']
            fill_pink = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            for i, missing in enumerate(ontbrekende_mask, start=2):
                if missing:
                    for cell in ws_prezero[i]:
                        cell.fill = fill_pink

        # Eind timer
        end_time = time.perf_counter()
        st.markdown(f"⏱️ Verwerkingstijd: {end_time - start_time:.2f} seconden")

        st.markdown("<div class='section-header'>📊 Resultaatoverzicht</div>", unsafe_allow_html=True)
        st.markdown(f"""
        - ✅ Bon aanwezig: **{resultaten.count('Bon aanwezig')}**
        - ⚖️ Gewicht verschilt: **{sum([1 for r in redenen if r != ''])}**
        - ❌ Geen bon aanwezig: **{resultaten.count('Geen bon aanwezig')}**
        - 🔁 PreZero-bonnen zonder match in Avalex: **{ontbrekende_mask.sum()}**
        """)

        st.success("✅ Verwerking voltooid.")
        st.download_button(
            "📥 Download resultaatbestand",
            data=output.getvalue(),
            file_name="LZP_resultaat.xlsx"
        )
    else:
        st.error(f"❌ Vereiste kolommen ontbreken. PreZero: {required_prezero}, Avalex: {required_avalex}")
